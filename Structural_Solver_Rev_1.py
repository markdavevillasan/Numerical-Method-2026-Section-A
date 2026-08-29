"""
3D Structural Frame Solver - Revision 1
=======================================

Purpose
-------
A standalone 3D frame solver based on the supplied Rev. 1 model data.

Model
-----
- Geometry: 6 m x 6 m x 6 m cube
- Nodes: 8
- Members: 12
- Bottom supports: Nodes 1-4, pinned
- 6 DOF per node: UX, UY, UZ, RX, RY, RZ
- Global axes: X and Z = lateral, Y = vertical
- Beta angles: base beams = 0 deg, roof beams = 0 deg, columns = 90 deg

Important
---------
The supplied data did not include actual member section properties or loads.
Therefore, the MATERIAL/SECTION values below are clearly marked as EXAMPLE
values and must be replaced with project values before using analysis results
for engineering decisions.

Dependencies
------------
numpy
openpyxl
matplotlib
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import math

import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("Install openpyxl first: pip install openpyxl") from exc

try:
    import matplotlib.pyplot as plt
except ImportError as exc:
    raise SystemExit("Install matplotlib first: pip install matplotlib") from exc


# ---------------------------------------------------------------------------
# PROJECT SETTINGS
# ---------------------------------------------------------------------------

REVISION = "Rev. 1"
OUTPUT_DIR = Path("Structural_Solver_Rev_1_Output")
OUTPUT_XLSX = OUTPUT_DIR / "Structural_Solver_Rev_1.xlsx"
OUTPUT_PNG = OUTPUT_DIR / "Structural_Diagram_Rev_1.png"

# Units used internally:
# length = m, force = N, stress = Pa
#
# EXAMPLE VALUES ONLY - REPLACE WITH PROJECT VALUES.
E = 200.0e9       # Pa
G = 80.0e9        # Pa
A = 0.0100        # m^2
IY = 8.333e-6     # m^4
IZ = 8.333e-6     # m^4
J = 1.667e-5      # m^4


# ---------------------------------------------------------------------------
# DATA CLASSES
# ---------------------------------------------------------------------------

@dataclass
class Node:
    node_id: int
    x: float
    y: float
    z: float
    support: str = "Free"

    def coordinates(self) -> np.ndarray:
        return np.array([self.x, self.y, self.z], dtype=float)


@dataclass
class Section:
    E: float
    G: float
    A: float
    Iy: float
    Iz: float
    J: float


@dataclass
class Member:
    member_id: int
    i_node: int
    j_node: int
    member_type: str
    beta_deg: float
    section: Section
    release_i: Tuple[str, ...] = field(default_factory=tuple)
    release_j: Tuple[str, ...] = field(default_factory=tuple)

    def length_and_direction(self, nodes: Dict[int, Node]) -> Tuple[float, np.ndarray]:
        p1 = nodes[self.i_node].coordinates()
        p2 = nodes[self.j_node].coordinates()
        vector = p2 - p1
        length = float(np.linalg.norm(vector))
        if length <= 0.0:
            raise ValueError(f"Member {self.member_id} has zero length.")
        return length, vector / length


# ---------------------------------------------------------------------------
# MODEL DEFINITION
# ---------------------------------------------------------------------------

def build_model() -> Tuple[Dict[int, Node], List[Member]]:
    """Create the supplied 6 m cube model."""

    L = 6.0

    nodes = {
        1: Node(1, 0.0, 0.0, 0.0, "Pinned"),
        2: Node(2, L,   0.0, 0.0, "Pinned"),
        3: Node(3, L,   0.0, L,   "Pinned"),
        4: Node(4, 0.0, 0.0, L,   "Pinned"),
        5: Node(5, 0.0, L,   0.0, "Free"),
        6: Node(6, L,   L,   0.0, "Free"),
        7: Node(7, L,   L,   L,   "Free"),
        8: Node(8, 0.0, L,   L,   "Free"),
    }

    section = Section(E=E, G=G, A=A, Iy=IY, Iz=IZ, J=J)

    members = [
        Member(1, 1, 2, "Base Beam", 0.0, section),
        Member(2, 2, 3, "Base Beam", 0.0, section),
        Member(3, 3, 4, "Base Beam", 0.0, section),
        Member(4, 4, 1, "Base Beam", 0.0, section),

        Member(5, 5, 6, "Roof Beam", 0.0, section),
        Member(6, 6, 7, "Roof Beam", 0.0, section),
        Member(7, 7, 8, "Roof Beam", 0.0, section),
        Member(8, 8, 5, "Roof Beam", 0.0, section),

        Member(9, 1, 5, "Column", 90.0, section),
        Member(10, 2, 6, "Column", 90.0, section),
        Member(11, 3, 7, "Column", 90.0, section),
        Member(12, 4, 8, "Column", 90.0, section),
    ]

    return nodes, members


# ---------------------------------------------------------------------------
# DOF AND SUPPORT HANDLING
# ---------------------------------------------------------------------------

DOF_NAMES = ("UX", "UY", "UZ", "RX", "RY", "RZ")


def node_dofs(node_id: int) -> np.ndarray:
    """Return the six global DOF numbers for a node, using 1-based numbering."""
    start = (node_id - 1) * 6 + 1
    return np.arange(start, start + 6, dtype=int)


def support_restraints(node: Node) -> np.ndarray:
    """
    Return six restraint flags.

    Pinned support:
        UX, UY, UZ restrained
        RX, RY, RZ active

    Free:
        all six active
    """
    if node.support.lower() == "pinned":
        return np.array([1, 1, 1, 0, 0, 0], dtype=int)
    return np.zeros(6, dtype=int)


def get_restrained_dofs(nodes: Dict[int, Node]) -> List[int]:
    restrained = []
    for node_id in sorted(nodes):
        dofs = node_dofs(node_id)
        flags = support_restraints(nodes[node_id])
        restrained.extend(dof for dof, flag in zip(dofs, flags) if flag)
    return restrained


# ---------------------------------------------------------------------------
# LOCAL AXES AND BETA ANGLE
# ---------------------------------------------------------------------------

def rotation_about_axis(axis: np.ndarray, angle_rad: float) -> np.ndarray:
    """Rodrigues rotation matrix about a unit vector."""
    axis = axis / np.linalg.norm(axis)
    x, y, z = axis
    c = math.cos(angle_rad)
    s = math.sin(angle_rad)
    C = 1.0 - c

    return np.array([
        [c + x*x*C,     x*y*C - z*s, x*z*C + y*s],
        [y*x*C + z*s,   c + y*y*C,   y*z*C - x*s],
        [z*x*C - y*s,   z*y*C + x*s, c + z*z*C],
    ])


def local_axes(member: Member, nodes: Dict[int, Node]) -> np.ndarray:
    """
    Calculate the member local axes as row vectors.

    Local x:
        Along the member from i-node to j-node.

    Reference local y:
        Global Y for horizontal members.
        Global Z for members that are nearly parallel to global Y.

    Beta:
        Rotates the reference local y axis about local x.
        The resulting local z completes a right-handed coordinate system.

    Returns
    -------
    axes : (3, 3) ndarray
        Row 0 = local x expressed in global coordinates
        Row 1 = local y expressed in global coordinates
        Row 2 = local z expressed in global coordinates
    """
    _, x_axis = member.length_and_direction(nodes)

    global_y = np.array([0.0, 1.0, 0.0])
    global_z = np.array([0.0, 0.0, 1.0])

    # Avoid a near-zero cross product when the member is vertical.
    if abs(np.dot(x_axis, global_y)) < 0.90:
        reference = global_y
    else:
        reference = global_z

    # Project reference vector into the plane normal to local x.
    y_axis = reference - np.dot(reference, x_axis) * x_axis
    y_norm = np.linalg.norm(y_axis)
    if y_norm < 1.0e-12:
        raise ValueError(f"Cannot construct local axes for member {member.member_id}.")
    y_axis /= y_norm

    z_axis = np.cross(x_axis, y_axis)
    z_axis /= np.linalg.norm(z_axis)

    # Apply beta rotation about local x.
    beta = math.radians(member.beta_deg)
    R_beta = rotation_about_axis(x_axis, beta)
    y_axis = R_beta @ y_axis
    z_axis = R_beta @ z_axis

    # Re-orthogonalize to reduce numerical drift.
    y_axis -= np.dot(y_axis, x_axis) * x_axis
    y_axis /= np.linalg.norm(y_axis)
    z_axis = np.cross(x_axis, y_axis)
    z_axis /= np.linalg.norm(z_axis)

    return np.vstack((x_axis, y_axis, z_axis))


# ---------------------------------------------------------------------------
# 3D FRAME ELEMENT STIFFNESS
# ---------------------------------------------------------------------------

def local_frame_stiffness(section: Section, length: float) -> np.ndarray:
    """
    Standard 3D Euler-Bernoulli frame element stiffness matrix.

    Local DOF order:
        [u, v, w, rx, ry, rz] at i
        [u, v, w, rx, ry, rz] at j
    """
    EA = section.E * section.A / length
    GJ = section.G * section.J / length

    EIy = section.E * section.Iy
    EIz = section.E * section.Iz

    k = np.zeros((12, 12), dtype=float)

    # Axial
    k[0, 0] = EA
    k[0, 6] = -EA
    k[6, 0] = -EA
    k[6, 6] = EA

    # Torsion
    k[3, 3] = GJ
    k[3, 9] = -GJ
    k[9, 3] = -GJ
    k[9, 9] = GJ

    # Bending about local z (v-rz)
    a = 12.0 * EIz / length**3
    b = 6.0 * EIz / length**2
    c = 4.0 * EIz / length
    d = 2.0 * EIz / length

    k[1, 1] += a
    k[1, 5] += b
    k[1, 7] += -a
    k[1, 11] += b

    k[5, 1] += b
    k[5, 5] += c
    k[5, 7] += -b
    k[5, 11] += d

    k[7, 1] += -a
    k[7, 5] += -b
    k[7, 7] += a
    k[7, 11] += -b

    k[11, 1] += b
    k[11, 5] += d
    k[11, 7] += -b
    k[11, 11] += c

    # Bending about local y (w-ry)
    a = 12.0 * EIy / length**3
    b = 6.0 * EIy / length**2
    c = 4.0 * EIy / length
    d = 2.0 * EIy / length

    k[2, 2] += a
    k[2, 4] += -b
    k[2, 8] += -a
    k[2, 10] += -b

    k[4, 2] += -b
    k[4, 4] += c
    k[4, 8] += b
    k[4, 10] += d

    k[8, 2] += -a
    k[8, 4] += b
    k[8, 8] += a
    k[8, 10] += b

    k[10, 2] += -b
    k[10, 4] += d
    k[10, 8] += b
    k[10, 10] += c

    return k


def transformation_matrix(axes: np.ndarray) -> np.ndarray:
    """
    Build the 12x12 local/global transformation matrix.

    axes rows are local x/y/z expressed in global coordinates.
    """
    R = axes
    T = np.zeros((12, 12), dtype=float)

    # Translation and rotation blocks use the same direction-cosine matrix.
    T[0:3, 0:3] = R
    T[3:6, 3:6] = R
    T[6:9, 6:9] = R
    T[9:12, 9:12] = R

    return T


def release_local_stiffness(
    k_local: np.ndarray,
    release_i: Tuple[str, ...],
    release_j: Tuple[str, ...],
) -> np.ndarray:
    """
    Apply approximate member-end releases by static condensation.

    Supported local DOF labels:
        UX, UY, UZ, RX, RY, RZ

    A released DOF is condensed from the element stiffness matrix. This is
    appropriate for idealized zero-stiffness end releases without adding
    artificial springs.
    """
    local_index = {name: i for i, name in enumerate(DOF_NAMES)}
    released = []

    for name in release_i:
        if name.upper() not in local_index:
            raise ValueError(f"Invalid i-end release: {name}")
        released.append(local_index[name.upper()])

    for name in release_j:
        if name.upper() not in local_index:
            raise ValueError(f"Invalid j-end release: {name}")
        released.append(6 + local_index[name.upper()])

    released = sorted(set(released))
    if not released:
        return k_local.copy()

    retained = [i for i in range(12) if i not in released]

    Krr = k_local[np.ix_(retained, retained)]
    Krp = k_local[np.ix_(retained, released)]
    Kpp = k_local[np.ix_(released, released)]

    # Condensation can fail for a torsional/axial release with no coupling.
    # In that case the released DOF is simply assigned zero stiffness.
    if np.linalg.matrix_rank(Kpp, tol=1.0e-12 * max(1.0, np.max(np.abs(Kpp)))) < len(released):
        k_out = np.zeros_like(k_local)
        k_out[np.ix_(retained, retained)] = Krr
        return k_out

    condensed = Krr - Krp @ np.linalg.solve(Kpp, Krp.T)

    k_out = np.zeros_like(k_local)
    k_out[np.ix_(retained, retained)] = condensed
    return k_out


# ---------------------------------------------------------------------------
# GLOBAL STIFFNESS ASSEMBLY
# ---------------------------------------------------------------------------

def member_global_stiffness(
    member: Member,
    nodes: Dict[int, Node],
) -> Tuple[np.ndarray, np.ndarray, float]:
    length, _ = member.length_and_direction(nodes)
    axes = local_axes(member, nodes)

    k_local = local_frame_stiffness(member.section, length)
    k_local = release_local_stiffness(
        k_local, member.release_i, member.release_j
    )

    T = transformation_matrix(axes)
    k_global = T.T @ k_local @ T

    return k_global, axes, length


def assemble_global_stiffness(
    nodes: Dict[int, Node],
    members: List[Member],
) -> Tuple[np.ndarray, Dict[int, np.ndarray], Dict[int, float]]:
    total_dof = len(nodes) * 6
    K = np.zeros((total_dof, total_dof), dtype=float)
    axes_by_member = {}
    lengths_by_member = {}

    for member in members:
        kg, axes, length = member_global_stiffness(member, nodes)

        i_dofs = node_dofs(member.i_node) - 1
        j_dofs = node_dofs(member.j_node) - 1
        edofs = np.concatenate((i_dofs, j_dofs))

        K[np.ix_(edofs, edofs)] += kg
        axes_by_member[member.member_id] = axes
        lengths_by_member[member.member_id] = length

    return K, axes_by_member, lengths_by_member


# ---------------------------------------------------------------------------
# LOADS AND SOLUTION
# ---------------------------------------------------------------------------

def build_load_vector(
    loads: Optional[Dict[int, Dict[str, float]]],
    number_of_nodes: int,
) -> np.ndarray:
    """Build the global nodal load vector from a dictionary."""
    F = np.zeros(number_of_nodes * 6, dtype=float)

    if not loads:
        return F

    dof_index = {name: i for i, name in enumerate(DOF_NAMES)}

    for node_id, node_loads in loads.items():
        if not 1 <= node_id <= number_of_nodes:
            raise ValueError(f"Load references invalid node {node_id}.")

        for dof_name, value in node_loads.items():
            key = dof_name.upper()
            if key not in dof_index:
                raise ValueError(f"Invalid load DOF: {dof_name}")
            F[(node_id - 1) * 6 + dof_index[key]] += float(value)

    return F


def solve_model(
    K: np.ndarray,
    F: np.ndarray,
    restrained_dofs_1based: List[int],
) -> Tuple[np.ndarray, np.ndarray]:
    """Solve K*u=F after imposing zero displacement at restrained DOFs."""
    n = len(F)
    restrained = np.array([d - 1 for d in restrained_dofs_1based], dtype=int)
    active = np.array([i for i in range(n) if i not in set(restrained)], dtype=int)

    if len(active) == 0:
        raise ValueError("There are no active DOFs.")

    Kaa = K[np.ix_(active, active)]
    Fa = F[active]

    rank = np.linalg.matrix_rank(Kaa, tol=1.0e-10 * max(1.0, np.max(np.abs(Kaa))))
    if rank < len(active):
        raise np.linalg.LinAlgError(
            "The active stiffness matrix is singular. "
            "Check supports, member releases, geometry, and section properties."
        )

    u = np.zeros(n, dtype=float)
    u[active] = np.linalg.solve(Kaa, Fa)

    reactions = K @ u - F
    return u, reactions


# ---------------------------------------------------------------------------
# MEMBER FORCE RECOVERY
# ---------------------------------------------------------------------------

def member_local_end_forces(
    member: Member,
    nodes: Dict[int, Node],
    global_displacements: np.ndarray,
) -> np.ndarray:
    """Recover member local end forces from global nodal displacements."""
    k_global, axes, _ = member_global_stiffness(member, nodes)
    T = transformation_matrix(axes)

    edofs = np.concatenate((
        node_dofs(member.i_node) - 1,
        node_dofs(member.j_node) - 1,
    ))
    u_global = global_displacements[edofs]
    u_local = T @ u_global

    k_local = local_frame_stiffness(
        member.section,
        member.length_and_direction(nodes)[0],
    )
    k_local = release_local_stiffness(
        k_local, member.release_i, member.release_j
    )

    return k_local @ u_local


# ---------------------------------------------------------------------------
# EXCEL OUTPUT
# ---------------------------------------------------------------------------

def write_excel(
    nodes: Dict[int, Node],
    members: List[Member],
    K: np.ndarray,
    loads: Dict[int, Dict[str, float]],
    displacements: Optional[np.ndarray],
    reactions: Optional[np.ndarray],
    axes_by_member: Dict[int, np.ndarray],
    lengths_by_member: Dict[int, float],
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Summary"

    summary = [
        ("Item", "Value"),
        ("Revision", REVISION),
        ("Cube edge length (m)", 6.0),
        ("Number of nodes", len(nodes)),
        ("Number of members", len(members)),
        ("Supported nodes", 4),
        ("Support type", "Pinned"),
        ("DOF per node", 6),
        ("Total DOF", len(nodes) * 6),
        ("Restrained DOF", len(get_restrained_dofs(nodes))),
        ("Active DOF", len(nodes) * 6 - len(get_restrained_dofs(nodes))),
        ("Global vertical axis", "Y"),
        ("Global lateral axes", "X and Z"),
        ("Beta angle, base beams (deg)", 0.0),
        ("Beta angle, roof beams (deg)", 0.0),
        ("Beta angle, columns (deg)", 90.0),
        ("Analysis status", "Solved" if displacements is not None else "Not solved - no loads"),
        ("Important note", "Section properties and loads must be replaced/defined for project analysis."),
    ]

    for row in summary:
        ws.append(row)

    # Nodes
    ws_nodes = wb.create_sheet("Nodes")
    ws_nodes.append(["Node", "X (m)", "Y (m)", "Z (m)", "Support"])
    for node_id in sorted(nodes):
        n = nodes[node_id]
        ws_nodes.append([n.node_id, n.x, n.y, n.z, n.support])

    # Node DOFs / restraints
    ws_dof = wb.create_sheet("Node DOFs")
    ws_dof.append([
        "Node", "X (m)", "Y (m)", "Z (m)", "Support",
        "Global DOF Range", "UX", "UY", "UZ", "RX", "RY", "RZ",
        "Active DOF"
    ])
    for node_id in sorted(nodes):
        n = nodes[node_id]
        dofs = node_dofs(node_id)
        flags = support_restraints(n)
        states = ["Restrained" if f else "Active" for f in flags]
        ws_dof.append([
            node_id, n.x, n.y, n.z, n.support,
            f"{dofs[0]} - {dofs[-1]}",
            *states,
            int(sum(1 for f in flags if not f)),
        ])

    # Supports
    ws_sup = wb.create_sheet("Restraints")
    ws_sup.append([
        "Node", "Support", "UX", "UY", "UZ", "RX", "RY", "RZ",
        "Restraint Code"
    ])
    for node_id in sorted(nodes):
        flags = support_restraints(nodes[node_id])
        code = "".join(str(int(x)) for x in flags)
        ws_sup.append([
            node_id,
            nodes[node_id].support,
            *["Restrained" if f else "Active" for f in flags],
            code,
        ])

    # Members
    ws_mem = wb.create_sheet("Members")
    ws_mem.append([
        "Member", "Node i (Start)", "Node j (End)", "Type",
        "Length (m)", "Beta (deg)",
        "i-End Releases", "j-End Releases"
    ])
    for m in members:
        ws_mem.append([
            m.member_id, m.i_node, m.j_node, m.member_type,
            lengths_by_member[m.member_id], m.beta_deg,
            ", ".join(m.release_i) if m.release_i else "None",
            ", ".join(m.release_j) if m.release_j else "None",
        ])

    # Local axes
    ws_axes = wb.create_sheet("Local Axes")
    ws_axes.append([
        "Member", "Node i", "Node j", "Type", "Beta (deg)", "Length (m)",
        "Local x-X", "Local x-Y", "Local x-Z",
        "Local y-X", "Local y-Y", "Local y-Z",
        "Local z-X", "Local z-Y", "Local z-Z"
    ])
    for m in members:
        a = axes_by_member[m.member_id]
        ws_axes.append([
            m.member_id, m.i_node, m.j_node, m.member_type,
            m.beta_deg, lengths_by_member[m.member_id],
            *a[0], *a[1], *a[2],
        ])

    # Releases
    ws_rel = wb.create_sheet("Releases")
    ws_rel.append([
        "Member", "Type", "i Node", "j Node",
        "i-End Release", "j-End Release"
    ])
    for m in members:
        ws_rel.append([
            m.member_id, m.member_type, m.i_node, m.j_node,
            ", ".join(m.release_i) if m.release_i else "None",
            ", ".join(m.release_j) if m.release_j else "None",
        ])

    # Loads
    ws_load = wb.create_sheet("Loads")
    ws_load.append(["Node", "DOF", "Load"])
    if loads:
        for node_id, node_loads in sorted(loads.items()):
            for dof, value in node_loads.items():
                ws_load.append([node_id, dof, value])
    else:
        ws_load.append(["-", "-", "No loads defined"])

    # Analysis results
    ws_res = wb.create_sheet("Analysis Results")
    ws_res.append([
        "Node", "UX", "UY", "UZ", "RX", "RY", "RZ",
        "RX/Radians", "RY/Radians", "RZ/Radians"
    ])

    if displacements is not None:
        for node_id in sorted(nodes):
            u = displacements[node_dofs(node_id) - 1]
            ws_res.append([node_id, *u])
    else:
        ws_res.append(["-", "No analysis performed", "", "", "", "", "", "", "", ""])

    # Reactions
    ws_react = wb.create_sheet("Support Reactions")
    ws_react.append(["Node", "FX", "FY", "FZ", "MX", "MY", "MZ"])
    if reactions is not None:
        for node_id in sorted(nodes):
            r = reactions[node_dofs(node_id) - 1]
            if nodes[node_id].support.lower() == "pinned":
                ws_react.append([node_id, *r])
    else:
        ws_react.append(["-", "No analysis performed", "", "", "", "", ""])

    # Member forces
    ws_forces = wb.create_sheet("Member Forces")
    ws_forces.append([
        "Member", "N_i", "Vy_i", "Vz_i", "T_i", "My_i", "Mz_i",
        "N_j", "Vy_j", "Vz_j", "T_j", "My_j", "Mz_j"
    ])

    if displacements is not None:
        for m in members:
            f = member_local_end_forces(m, nodes, displacements)
            ws_forces.append([m.member_id, *f])
    else:
        ws_forces.append(["-", "No analysis performed", "", "", "", "", "", "", "", "", "", "", ""])

    # Model matrix metadata
    ws_matrix = wb.create_sheet("Solver Info")
    ws_matrix.append(["Parameter", "Value"])
    ws_matrix.append(["Global stiffness matrix size", f"{K.shape[0]} x {K.shape[1]}"])
    ws_matrix.append(["DOF order", "[UX, UY, UZ, RX, RY, RZ]"])
    ws_matrix.append(["Support code", "1 = restrained, 0 = active"])
    ws_matrix.append(["Beta convention", "Positive rotation about local x-axis"])
    ws_matrix.append(["Units", "m, N, Pa, N-m"])
    ws_matrix.append([
        "Engineering note",
        "Verify section properties, loads, releases, units, and design assumptions before use."
    ])

    # Formatting
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            max_len = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            width = min(max(max_len + 2, 10), 45)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    # Simple header fill for readability.
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill

    wb.save(OUTPUT_XLSX)


# ---------------------------------------------------------------------------
# STRUCTURAL DIAGRAM
# ---------------------------------------------------------------------------

def plot_structure(
    nodes: Dict[int, Node],
    members: List[Member],
    axes_by_member: Dict[int, np.ndarray],
    lengths_by_member: Dict[int, float],
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    fig = plt.figure(figsize=(13, 9))
    ax = fig.add_subplot(111, projection="3d")

    # Members
    for m in members:
        p1 = nodes[m.i_node].coordinates()
        p2 = nodes[m.j_node].coordinates()

        if "column" in m.member_type.lower():
            line_width = 2.5
            line_style = "-"
        else:
            line_width = 2.2
            line_style = "-"

        ax.plot(
            [p1[0], p2[0]],
            [p1[1], p2[1]],
            [p1[2], p2[2]],
            linestyle=line_style,
            linewidth=line_width,
            label=m.member_type if m.member_id in (1, 9) else None,
        )

        mid = (p1 + p2) / 2.0
        ax.text(mid[0], mid[1], mid[2], f"M{m.member_id}", fontsize=8)

    # Nodes and DOF ranges
    for node_id in sorted(nodes):
        n = nodes[node_id]
        p = n.coordinates()
        ax.scatter([p[0]], [p[1]], [p[2]], s=35)

        dofs = node_dofs(node_id)
        ax.text(
            p[0] + 0.15, p[1] + 0.15, p[2] + 0.15,
            f"N{node_id}\nDOF {dofs[0]}-{dofs[-1]}",
            fontsize=8,
        )

    # Pinned support symbols as triangular markers below supported nodes.
    for node_id in sorted(nodes):
        n = nodes[node_id]
        if n.support.lower() == "pinned":
            ax.scatter(
                [n.x], [n.y - 0.20], [n.z],
                marker="^", s=120, depthshade=True
            )

    # Global axes
    origin = np.array([0.0, 0.0, 0.0])
    axis_len = 1.2
    global_axes = [
        (np.array([1.0, 0.0, 0.0]), "Global X"),
        (np.array([0.0, 1.0, 0.0]), "Global Y"),
        (np.array([0.0, 0.0, 1.0]), "Global Z"),
    ]

    for direction, label in global_axes:
        ax.quiver(
            *origin, *(direction * axis_len),
            arrow_length_ratio=0.12, linewidth=2.0
        )
        tip = direction * axis_len
        ax.text(tip[0], tip[1], tip[2], label, fontsize=9)

    # Local axes
    local_axis_len = 0.75
    for m in members:
        p1 = nodes[m.i_node].coordinates()
        p2 = nodes[m.j_node].coordinates()
        mid = (p1 + p2) / 2.0
        axes = axes_by_member[m.member_id]

        for idx, label in enumerate(("x", "y", "z")):
            d = axes[idx]
            ax.quiver(
                *mid, *(d * local_axis_len),
                arrow_length_ratio=0.16,
                linewidth=1.0,
            )
            end = mid + d * local_axis_len
            ax.text(end[0], end[1], end[2], f"l{label}", fontsize=7)

    ax.set_title(
        "6 m x 6 m x 6 m Cube - Structural Model, Rev. 1\n"
        "Pinned supports at nodes 1-4, member local axes and beta angles shown"
    )
    ax.set_xlabel("X (m) - lateral")
    ax.set_ylabel("Y (m) - vertical")
    ax.set_zlabel("Z (m) - lateral")

    # Equal aspect ratio.
    ax.set_box_aspect((1, 1, 1))

    # Model information panel.
    info = (
        "MODEL DATA - REV. 1\n\n"
        "Geometry\n"
        "  Cube edge       6.0 m\n"
        "  Nodes           8\n"
        "  Members         12\n"
        "  Vertical axis   Global Y\n\n"
        "Supports\n"
        "  Type            Pinned\n"
        "  Nodes           1, 2, 3, 4\n"
        "  Restrained      UX, UY, UZ\n\n"
        "Degrees of freedom\n"
        "  DOF per node    6\n"
        "  Total DOF       48\n"
        "  Restrained      12\n"
        "  Active          36\n\n"
        "Beta angles\n"
        "  Base beams      0 deg\n"
        "  Roof beams      0 deg\n"
        "  Columns         90 deg\n\n"
        "Local axes\n"
        "  x = member axis\n"
        "  y/z = transverse axes\n"
        "  beta rotates y/z about local x"
    )

    fig.text(
        0.74, 0.77, info,
        fontsize=8.5,
        va="top",
        bbox=dict(boxstyle="round,pad=0.5", alpha=0.12),
    )

    plt.tight_layout()
    fig.savefig(OUTPUT_PNG, dpi=200, bbox_inches="tight")
    plt.close(fig)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main() -> None:
    nodes, members = build_model()

    # -----------------------------------------------------------------------
    # LOAD INPUT
    # -----------------------------------------------------------------------
    # Add project loads here. Example:
    #
    # loads = {
    #     7: {"UY": -10000.0},
    # }
    #
    # The example is intentionally NOT activated because no project loads
    # were supplied.
    loads: Dict[int, Dict[str, float]] = {}

    K, axes_by_member, lengths_by_member = assemble_global_stiffness(
        nodes, members
    )

    restrained = get_restrained_dofs(nodes)
    F = build_load_vector(loads, len(nodes))

    displacements = None
    reactions = None

    if np.any(np.abs(F) > 0.0):
        displacements, reactions = solve_model(K, F, restrained)

    write_excel(
        nodes=nodes,
        members=members,
        K=K,
        loads=loads,
        displacements=displacements,
        reactions=reactions,
        axes_by_member=axes_by_member,
        lengths_by_member=lengths_by_member,
    )

    plot_structure(
        nodes=nodes,
        members=members,
        axes_by_member=axes_by_member,
        lengths_by_member=lengths_by_member,
    )

    print(f"{REVISION} solver completed.")
    print(f"Excel output: {OUTPUT_XLSX.resolve()}")
    print(f"Structural diagram: {OUTPUT_PNG.resolve()}")
    print(f"Total DOF: {len(nodes) * 6}")
    print(f"Restrained DOF: {len(restrained)}")
    print(f"Active DOF: {len(nodes) * 6 - len(restrained)}")
    print("No loads were supplied; therefore no displacement/reaction solution was performed.")


if __name__ == "__main__":
    main()
